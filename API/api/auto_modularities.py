"""
        NOTAS:

for root, dirs, files in os.walk(dir_path):
    for file in files: 
  
        # change the extension from '.mp3' to 
        # the one of your choice.
        if file.endswith('.mp3'):
            print (root+'/'+str(file))

#Para la caja TBLU se usan fusibles ATO con color claro por lo que al color se le agrega un "_clear", por ejemplo "ATO,10,red_clear"

#"F400": "ATO,15,BLUE
"""
           
from copy import copy
import requests
import openpyxl
import json
import os
from model import model
import gc
import pymysql
datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2=datos_conexion.datos_acceso()

modules = {}
modules_t = {}

fuses_types = {
    'PDC-P': {
        'MF1': "MULTI", 'MF2': "MULTI", 'F300': "ATO", 'F301': "MINI", 'F302': "MINI", 'F303': "MINI", 'F304': "MINI", 'F305': "MINI", 'F318': "MINI", 
        'F319': "MINI", 'F320': "MINI", 'F321': "MINI", 'F322': "MINI", 'F323': "MINI", 'F324': "MINI", 'F325': "MINI", 'F326': "ATO", 'F327': "ATO", 
        'F328': "ATO", 'F329': "ATO", 'F330': "ATO", 'F331': "ATO", 'F332': "ATO", 'F333': "ATO", 'F334': "ATO", 'F335': "ATO", 'E21': "CONN", 
        'E22': "CONN"
    },
    'PDC-D': {
        'F200': "MINI", 'F201': "MINI", 'F202': "MINI", 'F203': "MINI", 'F204': "MINI", 'F205': "MINI", 'F206': "MINI", 'F207': "MINI", 'F208': "MINI", 
        'F209': "ATO", 'F210': "ATO", 'F211': "ATO", 'F212': "ATO", 'F213': "ATO", 'F214': "ATO", 'F215': "ATO", 'F216': "ATO", 'F217': "MINI", 
        'F218': "MINI", 'F219': "MINI", 'F220': "MINI", 'F221': "MINI", 'F222': "MINI", 'F223': "MINI", 'F224': "MINI", 'F225': "MINI", 'F226': "MINI", 
        'F227': "MINI", 'F228': "MINI", 'F229': "MINI", 'F230': "MINI", 'F231': "MINI", 'F232': "MINI"
    },
    'PDC-R': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F449': "MAXI", 'F448': "MAXI", 'F447': "MAXI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 
        'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'F462': "MAXI", 
        'F463': "MAXI", 'F464': "MAXI", 'F465': "MINI", 'F466': "MINI", 'F467': "MINI", 'F468': "MINI", 'F469': "MINI", 'F470': "MINI", 'F471': "ATO", 
        'F472': "ATO", 'F473': "ATO", 'F474': "ATO", 'F475': "ATO", 'F476': "ATO", 'F477': "ATO", 'F478': "ATO", 'F479': "ATO", 'F480': "ATO", 
        'F481': "ATO", 'F482': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F96': "ATO"
    },
    'PDC-RMID': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-RS': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-S': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S9': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S21': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S17': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S19': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'PDC-S20': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'F96-1': {
        'F96': "ATO"
    },
    'F96': {
        'F96': "ATO"
    }, 
    'TBLU': {
        '9': "ATO", '8': "ATO", '7': "ATO", '6': "ATO", '5': "ATO", '4': "ATO", '3': "ATO", '2': "ATO", '1': "ATO"
    }
}

fuses_value = {
    'PDC-P': {
        'MF1': '', 'MF2': '', 'F300': '', 'F301': '', 'F302': '', 'F303': '', 'F304': '', 'F305': '', 'F318': '', 
        'F319': '', 'F320': '', 'F321': '', 'F322': '', 'F323': '', 'F324': '', 'F325': '', 'F326': '', 'F327': '', 
        'F328': '', 'F329': '', 'F330': '', 'F331': '', 'F332': '', 'F333': '', 'F334': '', 'F335': '', 'E21': '', 
        'E22': ''
    },
    'PDC-D': {
        'F200': '', 'F201': '', 'F202': '', 'F203': '', 'F204': '', 'F205': '', 'F206': '', 'F207': '', 'F208': '', 
        'F209': '', 'F210': '', 'F211': '', 'F212': '', 'F213': '', 'F214': '', 'F215': '', 'F216': '', 'F217': '', 
        'F218': '', 'F219': '', 'F220': '', 'F221': '', 'F222': '', 'F223': '', 'F224': '', 'F225': '', 'F226': '', 
        'F227': '', 'F228': '', 'F229': '', 'F230': '', 'F231': '', 'F232': ''
    },
    'PDC-R': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F449': '', 'F448': '', 'F447': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 
        'F454': '', 'F455': '', 'F456': '', 'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'F462': '', 
        'F463': '', 'F464': '', 'F465': '', 'F466': '', 'F467': '', 'F468': '', 'F469': '', 'F470': '', 'F471': '', 
        'F472': '', 'F473': '', 'F474': '', 'F475': '', 'F476': '', 'F477': '', 'F478': '', 'F479': '', 'F480': '', 
        'F481': '', 'F482': '', 'RELX': '', 'RELU': '', 'RELT': ''
    },
    'PDC-RMID': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 'F454': '', 'F455': '', 'F456': '', 
        'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'RELX': '', 'RELU': '', 'RELT': '', 'F449': '', 
        'F448': '', 'F447': '', 'F96': ''
    },
    'PDC-RS': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 'F454': '', 'F455': '', 'F456': '', 
        'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'RELX': '', 'RELU': '', 'RELT': '', 'F449': '', 
        'F448': '', 'F447': ''
    },
    'F96': {'F96': ''},    
    'PDC-S': {
        '1': '', '2': '', '3': '', '4': '', '5': '', '6': ''
    },
    'PDC-S9': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S21': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S17': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S19': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'PDC-S20': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'F96-1': {
        'F96': "ATO"
    },
    'TBLU': {
        '9': '', '8': '', '7': '', '6': '', '5': '', '4': '', '3': '', '2': '', '1': ''
    }
}

fuses_color = {
    #"1":    "negro", HMTEST ILX296270B1031517 EL.
    "5":    {"N000000008698":"beige", "N000000008708":"beige", "N000000004202":"beigeClear", "N000000006465":"beige"},
    "7.5":  {"N000000008699":"cafe", "N000000008709":"cafe", "N000000006466":"cafe"},
    "10":   {"N000000008700":"rojo", "N000000008710":"rojo", "N000000004204":"rojoClear"},
    "15":   {"N000000008701":"azul", "N000000008711":"azul"},
    "20":   {"N000000008702":"amarillo"},
    "25":   {"N000000008703":"natural"},
    "30":   {"N000000008704": "verde", "N000000007658":"verde"},
    "40":   {"N000000007659": "naranja"},
    "50":   {"N000000007660":"rojo"},
    "60":   {"A0009821923":"1008695"},
    "70":   {"A0025429419":"1010733"}
    # "60":   "azul"
    }

##################################### Modules management #################################
def refreshModules(db_event):
    """
    Ejecuta el flujo completo de actualización de módulos.

    La función genera la estructura de módulos a partir de los archivos
    Excel cargados y posteriormente sincroniza la información resultante
    para hacia las tablas de módulos de fusibles, torques
    y covers.

    Args:
        db_event (str): Identificador del evento en la base de datos (DBEVENT).

    Returns:
        None

    Raises:
        FileNotFoundError: Si la carpeta de módulos no existe.
        requests.exceptions.RequestException: Si ocurre un error durante
            la comunicación con la API.
        Exception: Cualquier error no controlado durante el procesamiento.
    """

    # Genera e inserta los modulos de fusibles en la base de datos
    fuses_data = makeFuseModules(db_event)
    updateFuseModules(fuses_data)
    
    # Genera e inserta los modulos de torque en la base de datos
    torques_data = makeTorqueModules(db_event)
    updateTorquesModules(torques_data)
    
    # Genera e inserta los modulos de covers en la base de datos
    covers_data = makeCoverModules(db_event)
    updateCoversModules(covers_data)
    
# def makeModules(data):
#     global modules, modules_t
#     #print("Data dentro de la creación de módulos!: ",data)
#     # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
#     endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
#     pdcrVariantes = requests.get(endpoint).json()
#     print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
#     modules = {}
#     modules_t = {}
#     print("#################### Modules ####################")
#     print("Modulos actual en MODULES: ",modules)
#     print("Modulos actual en MODULEST: ",modules_t)
#     dir_path = os.path.join(os.getcwd(), '..\\modules\\')
#     file_name = None
#     for root, dirs, files in os.walk(dir_path):
#         for file_name in files: 
#             print("file_name: ",file_name)
#             if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
#                 file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
#                 sheets = file.sheetnames
#                 for sheet in sheets: #para skipear las sheet que son correspondientes a información o X294 Izquierda F96
#                     if "Acomodos Modularidades" in sheet or "X294 Izquierda F96" in sheet:
#                         continue
#                     #para torques
#                     if "MFB" in sheet or "BATTERY" in sheet:
#                         currentSheet = file[sheet]
#                         for column in range(11, currentSheet.max_column + 1):
#                             module = currentSheet.cell(row = 3, column = column).value

#                             if isinstance(module,str):
#                                 module = module.replace(" ","")#se eliminan posibles espacios existentes, solo en str, porque puede haber valores None
#                             print("Modulo: ",module)

#                             if not(module in modules_t):
#                                 modules_t[module] = {}
#                             for row in range(5,currentSheet.max_row  + 1):
#                                 value = currentSheet.cell(row = row, column = column).value

#                                 if isinstance(value,str):
#                                     value = value.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                 if value == "x" or value == "X":
#                                     box = currentSheet.cell(row = row, column = 1).value
#                                     if isinstance(box,str):
#                                         box = box.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                     if box == "MFB-E":
#                                         print("Modulo de Nueva Caja: ",module)
#                                     torque = currentSheet.cell(row = row, column = 2).value
#                                     if isinstance(torque,str):
#                                         torque = torque.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                     if box == "MFB-S" and torque== "G1/21":
#                                         print("Aquí viene una battery-2",module)
#                                         box = "BATTERY-2"
#                                         torque = "BT"
#                                         print("Box Actualizado: ",box)
#                                         print("Torque Actualizado: ",torque)   
#                                     if box == "MFB-S2" and torque== "G1/21":
#                                         print("Aquí viene una battery-3",module,' : ',box)
#                                         box = "BATTERY-3"
#                                         torque = "BT"
#                                         print("Box Actualizado: ",box)
#                                         print("Torque Actualizado: ",torque)
#                                     if box == "BATTERY-3" and torque== "G1/21":
#                                         print("Aquí viene una battery-3",module,' : ',box)
#                                         torque ="BT"
#                                         print("Torque Actualizado: ",torque)
#                                     if not(box in modules_t[module]):
#                                         modules_t[module][box] = {}
#                                     modules_t[module][box][torque] = True
#                     #para fusibles
#                     else:
#                         currentSheet = file[sheet]
#                         for column in range(8, currentSheet.max_column + 1):
#                             module = currentSheet.cell(row = 3, column = column).value #se obtiene el valor de la celda que contiene el nombre del módulo

#                             if isinstance(module,str):
#                                 module = module.replace(" ","")#se eliminan posibles espacios existentes, solo en str, porque puede haber valores None
#                             print("Modulo: ",module)

#                             if not(module in modules):
#                                 modules[module] = {}
#                             for row in range(5,currentSheet.max_row  + 1):
#                                 value = currentSheet.cell(row = row, column = column).value #se obtiene el valor de la celda ej: "X"

#                                 if isinstance(value,str):
#                                     value = value.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                 if value == "x" or value == "X":
#                                     box = currentSheet.cell(row = row, column = 1).value.strip() #se obtiene el nombre de la caja primer columna de excel ej: PDC-P

#                                     if isinstance(box,str):
#                                         box = box.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                     if box =="Fuse Box F55" or box =="FuseBoxF55":
#                                         box = "TBLU"
#                                     fuse = currentSheet.cell(row = row, column = 2).value #ejemplo: F210
#                                     if isinstance(fuse,str):
#                                         fuse = fuse.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

#                                     if box == "TBLU":
#                                         fuse = fuse.replace("A", "")
#                                     if "PDC-S" in box:
#                                         fuse = str(fuse)
#                                         #print("Tipo del Fuse Ya convertido: ",type(fuse))
#                                     if "F96" in box:
#                                         print("Caja F96 AQUI",module)
#                                     if box == "PDC-R":
#                                         if module in pdcrVariantes["large"]:
#                                             box = "PDC-R"
#                                         elif module in pdcrVariantes["medium"]:
#                                             box = "PDC-RMID"
#                                         elif module in pdcrVariantes["small"]:
#                                             box = "PDC-RS"
#                                         else:
#                                             box = "PDC-RS"
#                                         #print("VARIANTE PARA PDC-R DEL MÓDULO: ",box)
#                                         if fuse == "X" or fuse == "T" or fuse == "U":
#                                             fuse = "REL" + fuse
#                                     amp = currentSheet.cell(row = row, column = 7).value #se obtiene el valor de la celda, ej: 7.
#                                     mercedes = currentSheet.cell(row = row, column = 4).value #se obtiene el valor de la celda, ej: N000000008708

#                                     if isinstance(amp,str):
#                                         amp = amp.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None
#                                     elif isinstance(amp,int):
#                                         amp = str(amp) + "A" #es int porque en el excel le faltó la letra A de amperes
#                                         print("____________________________________________")
#                                         print("revisar excel, se agregó una A en: ",amp)
#                                         print("____________________________________________")
#                                     elif isinstance(amp,float):
#                                         amp = str(amp) + "A" #es un float porque en el excel le faltó la letra A
#                                         print("____________________________________________")
#                                         print("revisar excel, se agregó una A en: ",amp)
#                                         print("____________________________________________")

#                                     if not(box in modules[module]):
#                                         modules[module][box] = {}
#                                     modules[module][box][fuse] = [amp[:-1], mercedes] #se quita la A de amp para que quede solo el número ej: 7
#                 del file
#                 gc.collect()
#                 os.remove(root+'\\'+ file_name)

#     structured_data = []
#     for module in modules:
#         temp = {
#             "DBEVENT": data,
#             "MODULO": "",
#             "CAJA_1": {},
#             "CAJA_2": {},
#             "CAJA_3": {},
#             "CAJA_4": {},
#             "CAJA_5": {},
#             "CAJA_6": {},
#             "CAJA_7": {},
#             "CAJA_8": {},
#             "CAJA_9": {},
#             "CAJA_10": {},
#             "CAJA_11": {},
#             "CAJA_12": {},
#             "CAJA_13": {},
#             "CAJA_14": {},
#             "CAJA_15": {},
#             "CAJA_16": {}
#             }

#         temp["MODULO"] = module
#         cnt = 1
#         for box in modules[module]:
#             try:
#                 if "PDC-P" in box:
#                     if not box in modules_t[module]:
#                         modules_t[module][box] = {"E1": True}
#                 if "PDC-D" in box:
#                     if not box in modules_t[module]:
#                         modules_t[module][box] = {"E1": True}
#                 if "PDC-R" in box:
#                     if not box in modules_t[module]:
#                         modules_t[module][box] = {"E1": True}
#             except Exception as ex:
#                 print (ex)
#                 #print("Inconcistencia de modulos:")
#                 #print("Caja: ", box)
#                 #print("Modulo de vision: ",module)
#                 #print("Modulo inexistente para torque")
#             key = "CAJA_" + str(cnt)
#             if not(box in temp[key]):
#                 temp[key][box]  = {}
#             for fuse in modules[module][box]:
#                 try:
#                     amp     = modules[module][box][fuse][0]
#                     mercedes     = modules[module][box][fuse][1]
#                     color = ""
#                     # if amp == "60":
#                     #     color = "1008695"
#                     # elif amp == "70":
#                     #     color = "1010733"
#                     # else:
#                     color   = fuses_color[amp][mercedes]
#                     temp[key][box][fuse] = color
#                 except Exception as ex:
#                     print("\n Vision exception in [", module, "] [", box, "] [", fuse, "]")
#                     print(ex)
#             cnt += 1
#         structured_data.append(temp)

#     print ("\n total de modulos vision: ",len(structured_data))

#     #print("modules_t:")
#     #print(modules_t)


#     torque_data = []
#     for module in modules_t:
#         temp = {
#             "DBEVENT": data,
#             "MODULO": "",
#             "CAJA_1": {},
#             "CAJA_2": {},
#             "CAJA_3": {},
#             "CAJA_4": {},
#             "CAJA_5": {},
#             "CAJA_6": {},
#             "CAJA_7": {},
#             "CAJA_8": {},
#             "CAJA_9": {},
#             "CAJA_10": {},
#             "CAJA_11": {},
#             "CAJA_12": {},
#             "CAJA_13": {},
#             "CAJA_14": {},
#             "CAJA_15": {},
#             "CAJA_16": {}
#             }

#         temp["MODULO"] = module
#         cnt = 1
#         for box in modules_t[module]:
#             key = "CAJA_" + str(cnt)
#             if not(box in temp[key]):
#                 temp[key][box]  = {}
#             try:
#                 for torque in modules_t[module][box]:
#                     try:
#                         temp[key][box][torque] = modules_t[module][box][torque]
#                     except Exception as ex:
#                         print("\nTorque exception in [", module, "] [", box, "] [", fuse, "]")
#                         print(ex)
#             except Exception as ex:
#                 print(ex)
#             cnt += 1
#         key = "CAJA_" + str(cnt)
#         if not("BATTERY" in temp[key]):
#             temp[key]["BATTERY"]  = {"BT": True}
#         torque_data.append(temp)

#     print ("\n total de modulos de torque: ",len(torque_data))

#     return structured_data, torque_data

def makeFuseModules(db_event):
    """
    Procesa la matriz de módulos y genera la estructura de fusibles
    asociada a cada módulo del evento.

    La función lee archivos Excel ubicados en el directorio ``modules``,
    identifica los módulos definidos en cada hoja válida y construye
    una estructura con la información de cajas, fusibles,
    amperajes y colores.

    Posteriormente, transforma la información en un formato serializable
    compatible con la tabla ``modulos_fusibles``.

    Args:
        db_event (str): Identificador del evento en la base de datos (DBEVENT).

    Returns:
        list[dict]: Lista de registros estructurados para inserción
        en la tabla ``modulos_fusibles``.

        Cada elemento contiene:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``FUSIBLES_DATA`` (str): Información serializada de fusibles
              agrupada por caja.

    Raises:
        FileNotFoundError: Si el directorio de módulos no existe.
        requests.exceptions.RequestException: Si falla la consulta de
            variantes PDC-R.
        pymysql.MySQLError: Si ocurre un error de conexión con MySQL.
        Exception: Cualquier error no controlado durante el procesamiento.
    """
    global modules
    
    # Consulta las variantes PDC-R cargadas en el evento
    endpoint = f"http://{host}:5000/api/get/{db_event}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    
    modules = {}
    print("#################### Fuses Modules ####################")
    print("Modulos anteriormente cargados: ",modules)
    
    dir_path = os.path.join(os.getcwd(), '..\\modules\\')
    file_name = None
    
    connection = pymysql.connect(host=host,user=user,password=password,database=db_event)
    cursor = connection.cursor()
    
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    # Omite hojas que no contienen información de módulos de fusibles
                    if "Acomodos Modularidades" in sheet or "MFB" in sheet or "BATTERY" in sheet or "Piezas" in sheet or "Cover" in sheet:
                        continue
                    currentSheet = file[sheet]
                    for column in range(8, currentSheet.max_column + 1):
                        
                        module = currentSheet.cell(row = 3, column = column).value
                        if module is None:
                            continue
                        
                        if not(module in modules):
                            # Inicializa la estructura del módulo
                            modules[module] = {}
                            #print("Modulo: ", module)
                        for row in range(5,currentSheet.max_row  + 1):
                            value = currentSheet.cell(row = row, column = column).value
                            if value == "x" or value == "X":
                                box = currentSheet.cell(row = row, column = 1).value
                                box = box.strip()
                                mercedes =  currentSheet.cell(row = row, column = 4).value
                                if box =="Fuse Box F55":
                                    box = "TBLU"
                                fuse = currentSheet.cell(row = row, column = 2).value
                                if box == "TBLU":
                                    fuse = fuse.replace("A", "")
                                if "PDC-S" in box or "F96" in box:
                                    fuse = str(fuse)
                                    fuse = fuse.strip()
                                # Determina la variante correspondiente de la caja PDC-R          
                                if box == "PDC-R":
                                    if module in pdcrVariantes["large"]:
                                        box = "PDC-R"
                                    elif module in pdcrVariantes["medium"]:
                                        box = "PDC-RMID"
                                    elif module in pdcrVariantes["small"]:
                                        box = "PDC-RS"
                                    else:
                                        box = "PDC-RS"
                                    if fuse == "X" or fuse == "T" or fuse == "U":
                                        fuse = "REL" + fuse
                                amp = currentSheet.cell(row = row, column = 7).value
                                amp = str(amp).strip()
                                
                                if not(box in modules[module]):
                                    modules[module][box] = {}
                                modules[module][box][fuse] = [amp[:-1], mercedes]
                                
                file.close()
                del file
                gc.collect()
    # Construye la estructura serializable requerida por modulos_fusibles
    structured_data = []
    for module in modules:
        fusibles_data = {}
        for box in modules[module]:
            # Inicializa la estructura de fusibles por caja
            if box not in fusibles_data:
                fusibles_data[box] = {}
            for fuse in modules[module][box]:
                try:
                    # Obtiene tipo, amperaje y color asociado al fusible
                    fuse_type    = fuses_types[box][fuse]
                    fuse_amp     = modules[module][box][fuse][0]
                    mercedes     = modules[module][box][fuse][1]
                    fuse_color = fuses_color[fuse_amp][mercedes]
                    
                    fusibles_data[box][fuse] = (
                        f"{fuse_type},{fuse_amp},{fuse_color}"
                    )
                except Exception as ex:
                    print("\nexception in [",module,"] [",box,"] [",fuse,"]")
                    print(ex)
                    
        structured_data.append({
            "DBEVENT": db_event,
            "MODULO": module,
            "FUSIBLES_DATA": json.dumps(fusibles_data)
        })

    print ("\n total de modulos: ",len(structured_data))
    return structured_data

def updateFuseModules(fuses_data):
    """
    Carga los módulos de fusibles en la tabla ``modulos_fusibles``.

    La función elimina previamente los registros existentes asociados
    a los módulos recibidos y posteriormente inserta la información
    actualizada generada desde la matriz de módulos.

    Args:
        modules_data (list[dict]): Lista de registros estructurados
        de módulos de fusibles.

        Cada elemento debe contener:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``FUSIBLES_DATA`` (str): Información serializada de fusibles.

    Returns:
        None

    Raises:
        pymysql.MySQLError: Si ocurre un error durante las operaciones
            de inserción o eliminación en la base de datos.
        IndexError: Si la lista recibida está vacía.
        Exception: Cualquier error no controlado durante el proceso.
    """
    print("updating fuse modules")

    if len(fuses_data) == 0:
        print("No hay datos para insertar")
        return

    connection = pymysql.connect(host=host,user=user,password=password,database=fuses_data[0]["DBEVENT"])
    cursor = connection.cursor()

    try:
        # Obtiene los módulos que deben ser reemplazados
        modulos = set()

        for item in fuses_data:
            modulos.add(item["MODULO"])

        # Elimina los registros existentes antes de insertar la nueva información
        for modulo in modulos:
            cursor.execute("""
                DELETE FROM modulos_fusibles
                WHERE MODULO=%s
            """, (modulo,))

        connection.commit()
        
        # Inserta la estructura actualizada de fusibles por módulo
        for record in fuses_data:
            modulo = record["MODULO"]
            fuse_data = record["FUSIBLES_DATA"]

            cursor.execute("""
                INSERT INTO modulos_fusibles
                (MODULO, FUSIBLES_DATA)
                VALUES (%s, %s)
            """, (modulo, fuse_data))
            
        # Confirma la carga de módulos en la base de datos
        connection.commit()

        print("Modulos fusibles actualizados correctamente")

    except Exception as ex:
        print("updateFuseModules Exception:", ex)

    finally:
        connection.close()

def makeTorqueModules(db_event):
    """
    Procesa la matriz de módulos y genera la estructura de torques
    asociada a cada módulo del evento.

    La función lee archivos Excel ubicados en el directorio ``modules``,
    identifica los módulos definidos en cada hoja válida y construye
    una estructura con la información de cajas y torques.

    Posteriormente, transforma la información en un formato serializable
    compatible con la tabla ``modulos_torques``.

    Args:
        db_event (str): Identificador del evento en la base de datos (DBEVENT).

    Returns:
        list[dict]: Lista de registros estructurados para inserción
        en la tabla ``modulos_torques``.

        Cada elemento contiene:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``TORQUES_DATA`` (str): Información serializada de torques
              agrupada por caja.

    Raises:
        FileNotFoundError: Si el directorio de módulos no existe.
        Exception: Cualquier error no controlado durante el procesamiento.
    """
    modules_t = {}
    print("#################### Torques Modules ####################")

    dir_path = os.path.join(os.getcwd(), '..\\modules\\')

    for root, dirs, files in os.walk(dir_path):
        for file_name in files:
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames

                for sheet in sheets:
                    # Omite hojas que no contienen información de módulos de torque
                    if not ("MFB" in sheet or "BATTERY" in sheet):
                        continue

                    currentSheet = file[sheet]

                    for column in range(11, currentSheet.max_column + 1):

                        module = currentSheet.cell(row=3, column=column).value
        
                        if module is None:
                            continue
                        
                        if isinstance(module,str):
                            module = module.replace(" ","")

                        if module not in modules_t:
                            modules_t[module] = {}

                        for row in range(5, currentSheet.max_row + 1):
                            value = currentSheet.cell(row=row, column=column).value
                            
                            if isinstance(value,str):
                                value = value.replace(" ","")

                            if value == "x" or value == "X":
                                box = currentSheet.cell(row=row, column=1).value
                                torque = currentSheet.cell(row=row, column=2).value

                                if box is None or torque is None:
                                    continue

                                box = str(box).replace(" ", "").strip()
                                torque = str(torque).replace(" ", "").strip()

                                # Reasignacion de caja y torques.
                                if box == "MFB-S" and torque == "G1/21":
                                    box = "BATTERY-2"
                                    torque = "BT"

                                elif box == "MFB-S2" and torque == "G1/21":
                                    box = "BATTERY-3"
                                    torque = "BT"

                                elif box == "MFB-S2" and torque != "G1/21":
                                    box = "MFB-P1"

                                if box not in modules_t[module]:
                                    modules_t[module][box] = {}

                                modules_t[module][box][torque] = True

                file.close()
                del file
                gc.collect()
    
    # Construye la estructura serializable requerida por modulos_torques
    structured_torque = []
    for module in modules_t:
        torques_data = {}
        for box in modules_t[module]:
            # Inicializa la estructura de torques por caja
            if box not in torques_data:
                torques_data[box] = {}
            for torque in modules_t[module][box]:
                torques_data[box][torque] = modules_t[module][box][torque]

        if len(torques_data) == 0:
            continue

        structured_torque.append({
            "DBEVENT": db_event,
            "MODULO": module,
            "TORQUES_DATA": json.dumps(torques_data),
            "ACTIVO": 1
        })

    print("\nTotal módulos torque:", len(structured_torque))
    print("\nEjemplo torque:")
    print(json.dumps(structured_torque[:5], indent=2))

    return structured_torque

def updateTorquesModules(torques_data):
    """
    Carga los módulos de torques en la tabla ``modulos_torques``.

    La función elimina previamente los registros existentes asociados
    a los módulos recibidos y posteriormente inserta la información
    actualizada generada desde la matriz de módulos.

    Args:
        torque_data (list[dict]): Lista de registros estructurados
        de módulos de torques.

        Cada elemento debe contener:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``TORQUES_DATA`` (str): Información serializada de torques.
            - ``ACTIVO`` (int): Estado lógico del registro.

    Returns:
        None

    Raises:
        pymysql.MySQLError: Si ocurre un error durante las operaciones
            de inserción o eliminación en la base de datos.
        IndexError: Si la lista recibida está vacía.
        Exception: Cualquier error no controlado durante el proceso.
    """
    print("updating torques")

    if len(torques_data) == 0:
        print("No hay torques para insertar")
        return

    connection = pymysql.connect(host=host,user=user,password=password,database=torques_data[0]["DBEVENT"])
    cursor = connection.cursor()

    try:
        # Elimina e inserta la información actualizada por módulo
        for record in torques_data:
            modulo = record["MODULO"]
            torque_data = record["TORQUES_DATA"]
            activo = record["ACTIVO"]

            cursor.execute("""
                DELETE FROM modulos_torques
                WHERE MODULO=%s
            """, (modulo,))

            cursor.execute("""
                INSERT INTO modulos_torques
                (MODULO, TORQUES_DATA, ACTIVO)
                VALUES (%s, %s, %s)
            """, (modulo, torque_data, activo))

        connection.commit()

        print("Torques actualizados correctamente")

    except Exception as ex:
        print("updateTorquesModules Exception:", ex)

    finally:
        connection.close()
        
def makeCoverModules(db_event):
    """
    Procesa la matriz de módulos y genera la estructura de covers
    asociada a cada módulo del evento.

    La función lee archivos Excel ubicados en el directorio ``modules``,
    identifica las hojas relacionadas con covers y construye una estructura
    que relaciona cada módulo con sus cajas y covers correspondientes.

    Posteriormente, transforma la información en un formato serializable
    compatible con la tabla ``modulos_covers``.

    Args:
        db_event (str): Identificador del evento en la base de datos (DBEVENT).

    Returns:
        list[dict]: Lista de registros estructurados para inserción
        en la tabla ``modulos_covers``.

        Cada elemento contiene:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``COVERS_DATA`` (str): Información serializada de covers.

    Raises:
        FileNotFoundError: Si el directorio de módulos no existe.
        Exception: Cualquier error no controlado durante el procesamiento.
    """
    print("#################### Covers Modules ####################")
    modules_cover = {}

    dir_path = os.path.join(os.getcwd(), '..\\modules\\')
    
    for root, dirs, files in os.walk(dir_path):
        for file_name in files:
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames

                for sheet in sheets:
                    # Omite hojas que no contienen información de módulos de covers
                    if "Cover" not in sheet:
                        continue

                    currentSheet = file[sheet]

                    for column in range(8, currentSheet.max_column + 1):

                        module = currentSheet.cell(row=3,column=column).value

                        if module is None:
                            continue

                        if isinstance(module,str):
                            module = module.replace(" ","")
                        
                        if module not in modules_cover:
                            modules_cover[module] = {}

                        for row in range(5, currentSheet.max_row + 1):

                            value = currentSheet.cell(row=row,column=column).value

                            if value == "x" or value == "X":
                                box = currentSheet.cell(row=row,column=1).value
                                cover = currentSheet.cell(row=row,column=2).value

                                if box is None or cover is None:
                                    continue

                                box = str(box).strip()
                                cover = str(cover).strip()

                                modules_cover[module][box] = cover

            file.close()
            del file
            gc.collect()
            os.remove(root + '\\' + file_name)

    # Construye la estructura serializable requerida por modulos_covers
    structured_covers = []
    for modulo in modules_cover:
        covers_data = modules_cover[modulo]

        if len(covers_data) == 0:
            continue

        structured_covers.append({
            "DBEVENT": db_event,
            "MODULO": modulo,
            "COVERS_DATA": json.dumps(covers_data)
        })

    print("\nTotal módulos covers:", len(structured_covers))

    print("\nEjemplo covers:")
    print(json.dumps(structured_covers[:5], indent=2))

    return structured_covers

def updateCoversModules(covers_data):
    """
    Sincroniza los módulos de covers contra la tabla ``modulos_covers``.

    La función elimina previamente los registros existentes asociados
    a los módulos recibidos y posteriormente inserta la información
    actualizada generada desde la matriz de módulos.

    Args:
        covers_data (list[dict]): Lista de registros estructurados
        de módulos de covers.

        Cada elemento debe contener:

            - ``DBEVENT`` (str): Identificador del evento.
            - ``MODULO`` (str): Identificador único del módulo.
            - ``COVERS_DATA`` (str): Información serializada de covers.

    Returns:
        None

    Raises:
        pymysql.MySQLError: Si ocurre un error durante las operaciones
            de inserción o eliminación en la base de datos.
        IndexError: Si la lista recibida está vacía.
        Exception: Cualquier error no controlado durante el proceso.
    """
    print("updating covers")

    if len(covers_data) == 0:
        print("No hay covers para insertar")
        return

    connection = pymysql.connect(host=host,user=user,password=password,database=covers_data[0]["DBEVENT"])
    cursor = connection.cursor()

    try:
        # Elimina e inserta la información actualizada por módulo
        for record in covers_data:
            modulo = record["MODULO"]
            cover_data = record["COVERS_DATA"]

            cursor.execute("""
                DELETE FROM modulos_covers
                WHERE MODULO=%s
            """, (modulo,))

            cursor.execute("""
                INSERT INTO modulos_covers
                (MODULO, COVERS_DATA)
                VALUES (%s, %s)
            """, (modulo, cover_data))

        connection.commit()

        print("Covers actualizados correctamente")

    except Exception as ex:
        print("updateCoversModules error:", ex)

    finally:
        connection.close()

def visionUpdate(data):
    print("vision updating")
    tabla = data[0]["DBEVENT"]
    #print("TABLAAAAA Vision: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/modulos_fusibles/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                #print(type(i["MODULO"])) es un string
                i["MODULO"] =  i["MODULO"].strip()
                endpoint = f"http://{host}:5000/api/post/modulos_fusibles"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modulos_fusibles/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def torqueUpdate(data):
    print("torque updating")
    tabla = data[0]["DBEVENT"]
    #print("TABLAAAAA Torque: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/modulos_torques/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                i["MODULO"] =  i["MODULO"].strip()
                endpoint = f"http://{host}:5000/api/post/modulos_torques"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modulos_torques/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def pdcrVariants (data):
    """

            IN CONSTRUCTION


    PDC-R small:  A2239060902
    PDC-R MEDIUM:  A2239061002
    PDC-R LARGE:  A2239061102
    """
    print("#################### pdcrVariants ####################")
    dir_path = os.path.join(os.getcwd(), '..\\FAAJISPREV\\')
    file_name = None
    rows = []
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if temp.endswith('.txt'):
                fic = open(dir_path + file_name)
                lines = list(fic)
                for i in lines:
                    i = i[:-1]
                    rows.append(i.split())
                print(len(lines))
                for i in range(5):
                    print(lines[i])


################################### Modularities management ##############################
def makeModularities(data):
    global modules
    print("Dentro de MakeModularities DATA: ",data)
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R:\n",pdcrVariantes)
    print("#################### Modularities ####################")
    endpoint = f"http://{host}:5000/api/get/{data}/modulos_fusibles/all/-/-/-/-/-"
    modulesExisting = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos VISION: ",modulesExisting["MODULO"])
    print("LEN VISION: ",len(modulesExisting["MODULO"]))

    endpoint = f"http://{host}:5000/api/get/{data}/modulos_torques/all/-/-/-/-/-"
    modulesExisting_t = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos TORQUES: ",modulesExisting_t["MODULO"])
    print("LEN TORQUES: ",len(modulesExisting_t["MODULO"]))

    dir_path = os.path.join(os.getcwd(), '..\\ILX\\')
    file_name = None
    modularities = []
    modulosFaltantes = []
    ilxfaltantes = {
        "ILX": {},
        "Modulos": {}
        }
    flujo = ""
    numero = ""
    if 'izquierda' in data:
        print('EVENTO DE CONDUCCION IZQUIERDA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'ILZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'ILX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'ILX'
            numero = '294'
    if 'derecha' in data:
        print('EVENTO DE CONDUCCION DERECHA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'IRZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'IRX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'IRX'
            numero = '294'        
    flujo_numero = flujo + numero
    
    global modules
    print("Dentro de MakeModularities DATA: ",data)
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R:\n",pdcrVariantes)
    print("#################### Modularities ####################")
    print("se obtienen los módulos existentes de fusibles cargados para este evento (obtenidos de la matriz)")
    endpoint = f"http://{host}:5000/api/get/{data}/modulos_fusibles/all/-/-/-/-/-"
    modulesExisting = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos VISION: ",modulesExisting["MODULO"])
    print("LEN VISION: ",len(modulesExisting["MODULO"]))

    endpoint = f"http://{host}:5000/api/get/{data}/modulos_torques/all/-/-/-/-/-"
    modulesExisting_t = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos TORQUES: ",modulesExisting_t["MODULO"])
    print("LEN TORQUES: ",len(modulesExisting_t["MODULO"]))
    dir_path = os.path.join(os.getcwd(), '..\\ILX\\')
    file_name = None
    modularities = []
    modulosFaltantes = []
    ilxfaltantes = {
        "ILX": {},
        "Modulos": {}
        }
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: #se recorre cada archivo existente dentro de el folder en que se cargaron
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper() #se obtiene el nombre de la modularidad separando el filename por el punto y quedando con la parte antes del punto en mayúsculas

            if not(flujo_numero in file_name):# SI NO se encuentra el nombre esperado de inicio para un arnés de este tipo:
                ilxfaltantes["ILX"][ILX] = {
                            "vision": [],
                            "torque": []
                            } #se crea un diccionario para esta modularidad
                ilxfaltantes["ILX"][ILX]["vision"].append("No es un DAT válido para este evento") #se agrega el mensaje que no es un DAT válido
                ilxfaltantes["ILX"][ILX]["torque"].append("No es un DAT válido para este evento") #se agrega el mensaje que no es un DAT válido
                modulosFaltantes.append(ILX) #se agrega a la lista final de módulos faltantes para que aparezca en pantalla
                ilxfaltantes["Modulos"] = modulosFaltantes #se actualiza esta lista
                os.remove(root+'\\'+ file_name) #se elimina el archivo de los DATS
            else:

                if temp.endswith('.dat'):
                    flag_s = False
                    flag_m = False
                    flag_l = False
                    qr_pdcr = {}
                    flag_mfbp2_der = False
                    flag_mfbp2 = []
                    fic = open(dir_path + file_name)
                    lines = list(fic)
                    csv = ""
                    for line in lines:
                        csv += line.rsplit(sep = "=")[-1][:-1] + ","
                    csv = csv[:-1]
                    fic.close()
                    #print("MODULOS DEL ILX: ",csv.split(sep = ","))
                    if "ILX294" in ILX:
                        print("Evento 294 IZQUIERDA")
                    if "IRX294" in ILX:
                        print("Evento 294 DERECHA")
                    if "ILX296" in ILX:
                        print("Evento 296 IZQUIERDA")
                    if "IRX296" in ILX:
                        print("Evento 296 DERECHA")
                    if "296" in ILX or "294" in ILX:
                        #print("Evento 296")
                        if "IRX" in ILX or "IRZ" in ILX:
                            print("Lleva la MFB-P2 DERECHA con terminación : 7216")
                            flag_mfbp2_der = True
                        for mod in csv.split(sep = ","):
                            if mod in pdcrVariantes["large"]:
                                flag_l = True
                            if mod in pdcrVariantes["medium"]:
                                flag_m = True
                            if mod in pdcrVariantes["small"]:
                                flag_s = True
                        print("\t\t+++++++++++ FLAGS de",ILX,":+++++++++++\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l," Flag MFB-P2 DER: ",flag_mfbp2_der)
                        if flag_mfbp2_der == True:
                            flag_mfbp2 = ["12975407216", True]
                        else:
                            flag_mfbp2 = ["12975407316", True]
                        if flag_l == True:
                            qr_pdcr = {
                                "PDC-R": ["12239061602", True],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-S2": ["12975402001", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                            print("Variante de caja PDC-R")
                        if flag_m == True and flag_l == False:
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["12239061502", True],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-S2": ["12975402001", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                            print("Variante de caja PDC-RMID")
                        if flag_s == True and flag_m == False:
                            print("Variante de caja PDC-RS")
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["12239061402", True],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-S2": ["12975402001", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                        if flag_s == False and flag_m == False and flag_l == False:
                            print("La caja no contiene módulos pertenecientes a las categorías.")
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-S2": ["12975402001", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }

                    temp = {
                        "DBEVENT": data,
                        "PEDIDO": ILX,
                        "DATETIME": "AUTO",
                        "MODULOS_VISION": {"INTERIOR": csv.split(sep = ",")},
                        "MODULOS_TORQUE": {"INTERIOR": csv.split(sep = ",")},
                        "MODULOS_ALTURA": {"INTERIOR": csv.split(sep = ",")},
                        "QR_BOXES": qr_pdcr,
                        "ACTIVE": 1
                        }
                    print("Códigos QR FINAL: ",qr_pdcr)
                    #print("ILX: ",ILX)
                    #print("Modulos que tiene: ",csv)
                    #print("Modulos que tiene TIPO: ",type(csv))
                    #print("Modulos que tiene el ILX: ",csv.split(","))
                    #print("Modulos que tiene convertido a array TIPO: ",type(csv.split(",")))
                    modulosDesconocidos = set(csv.split(",")) - set(modulesExisting["MODULO"])
                    modulosDesconocidos_t = set(csv.split(",")) - set(modulesExisting_t["MODULO"])
                    #print("Comparación; Modulos del ILX que NO están en la base de datos: ", modulosDesconocidos)
                    #print("Comparación; Modulos del ILX que NO están en la base de datos LEN VISION: ", len(modulosDesconocidos))
                    #print("Comparación; Modulos del ILX que NO están en la base de datos LEN TORQUES: ", len(modulosDesconocidos_t))
                    #print("Comparación tipo", type(modulosDesconocidos))
                    if len(modulosDesconocidos) == 0 and len(modulosDesconocidos_t) == 0:
                        modularities.append(temp)
                    else:
                        ilxfaltantes["ILX"][ILX] = {
                            "vision": [],
                            "torque": []
                            }
                        for e in modulosDesconocidos:
                            ilxfaltantes["ILX"][ILX]["vision"].append(e)
                        #print(e)
                            if not(e in modulosFaltantes):
                                modulosFaltantes.append(e)
                        for t in modulosDesconocidos_t:
                            #print(t)
                            ilxfaltantes["ILX"][ILX]["torque"].append(t)
                            if not(t in modulosFaltantes):
                                modulosFaltantes.append(t)
                    
                    ilxfaltantes["Modulos"] = modulosFaltantes
                    os.remove(root+'\\'+ file_name)
    #print("Lista total de Módulos Faltantes: ",ilxfaltantes)
    if len(modularities) != 0:
        updateModularities(modularities)
    return ilxfaltantes
 
def updateModularities(data):
    print("updating")
    #print("Data dentro de Upload Modularities: ",data)
    tabla = data[0]["DBEVENT"]
    print("TABLA en updating para DATS: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/pedidos/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("PEDIDO" in existing):
        existing["PEDIDO"] = []
    for i in data:
        try:
            if not(i["PEDIDO"] in existing["PEDIDO"]):
                endpoint = f"http://{host}:5000/api/post/pedidos"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["PEDIDO"].index(i["PEDIDO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/pedidos/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

##################################### Determinantes management #################################
def refreshDeterminantes(db_event, usuario):
    """
    Ejecuta el flujo completo de actualización de módulos determinantes.

    La función genera la estructura de módulos determinantes a partir
    de archivos Excel locales y posteriormente sincroniza los registros
    resultantes contra la tabla ``determinantes``.

    Args:
        db_event (str): Identificador del evento en la base de datos.
        usuario (str):  Identificador del usuario que ejecuta la carga.

    Returns:
        None

    Raises:
        IndexError: Si el procesamiento no genera registros válidos.
        requests.exceptions.RequestException: Si ocurre un error durante
            la comunicación con la API.
        FileNotFoundError: Si la carpeta de archivos Excel no existe.
    """

    print("#################### refreshDeterminantes ####################")

    determinantes_data = makeDeterminantes(db_event, usuario)
    updateDeterminantes(determinantes_data)

def makeDeterminantes(data,usuario):
    
    """
    Lee archivos Excel del directorio de determinantes, clasifica los módulos
    por variante de la PDC-R y genera una lista estructurada para su procesamiento.

    La función busca recursivamente archivos ``.xls`` y ``.xlsx`` en la carpeta
    ``../determinantes``, extrae la información de las hojas válidas, omite la
    hoja ``PIEZAS`` y organiza los módulos según las columnas correspondientes.
    Al finalizar, elimina los archivos procesados del directorio.

    Args:
        data (str): Identificador del evento en la base de datos.
        usuario (str): Identificador del usuario que ejecuta la carga.

    Returns:
        list[dict]: Lista de diccionarios con la información estructurada
        de los módulos determinantes. Cada elemento contiene:

            - ``DBEVENT`` (str): Identificador del evento en la base de datos.
            - ``MODULO`` (str): Módulo identificador.
            - ``VARIANTE`` (str): Tipo de caja PDC-R.
            - ``DATETIME`` (str): Marcador automático para asignación de timestamp.
            - ``USUARIO`` (str): Identificador del usuario que ejecuta la carga.
            - ``ACTIVO`` (int): Estado lógico del registro 0/1.

    Raises:
        FileNotFoundError: Si la ruta del directorio de determinantes no existe.
        PermissionError: Si el sistema no puede eliminar el archivo Excel procesado.
    """

    print("#################### makeDeterminantes ####################")

    global determinantes
    # Se crea un diccionario de determinantes para identificar el tipo de caja PDC-R
    determinantes = {
        "PDC-RS":[],
        "PDC-RMID":[],
        "PDC-R":[]
        }
    
    print("Modulos anteriormente cargados: ",determinantes)
    print("DATA que se pasa como arugmento a Determinantes",data)
    print("USUARIO que se pasa como arugmento a Determinantes",usuario)
    
    # Se obtiene la ruta de la carpeta "determinantes" ubicada un nivel arriba del directorio actual
    dir_path = os.path.join(os.getcwd(), '..\\determinantes\\')
    file_name = None
    
    # Se recorre recursivamente el contenido del directorio obtenido en dir_path
    for root, dirs, files in os.walk(dir_path):
        for file_name in files:
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    # La hoja PIEZAS no contiene módulos determinantes
                    if "PIEZAS" in sheet:
                        continue
                    currentSheet = file[sheet]
                    columnas = ["PDC-RS","PDC-RMID","PDC-R"]
                    for variante in columnas:
                        if variante == "PDC-RS":
                            col = 2
                        if variante == "PDC-RMID":
                            col = 5
                        if variante == "PDC-R":
                            col = 8
                        # Se inicia desde la fila 3 porque las primeras filas contienen encabezados
                        for row in range(3, currentSheet.max_row + 1): 
                            module = currentSheet.cell(column = col, row = row).value
                            
                            if not(module in determinantes[variante]):
                                if module != None:
                                    determinantes[variante].append(module)
                print("Arreglo final de determinantes: ",determinantes)
                del file
                gc.collect()
                os.remove(root+'\\'+ file_name)
    
    # Estructura la información en un formato plano compatible con el modelo
    # de la base de datos (tabla: ``determinantes``)
    structured_data = []
    for variante in determinantes:
        print("Variante: ",variante)
        for module in determinantes[variante]:
            print("Modulo: ",module)

            temp = {
            "DBEVENT": data,
            "MODULO": module,
            "VARIANTE": variante,
            "DATETIME": "AUTO",
            "USUARIO": usuario,
            "ACTIVO": 1
            }
            structured_data.append(temp)

    print ("\n total de modulos: ",len(structured_data))

    return structured_data

def updateDeterminantes(determinantes_data):
    """
    Sincroniza los módulos determinantes contra la tabla ``determinantes``.

    La función consulta los registros existentes asociados al evento recibido
    (``DBEVENT``) y determina si cada módulo debe insertarse o actualizarse
    mediante peticiones HTTP a la API.

    Args:
        determinantes_data (list[dict]): Lista de registros estructurados
        de módulos determinantes. Cada elemento debe contener los campos:

            - ``DBEVENT`` (str): Identificador del evento en la base de datos.
            - ``MODULO`` (str): Módulo identificador.
            - ``VARIANTE`` (str): Tipo de caja PDC-R.
            - ``DATETIME`` (str): Marcador automático para asignación de timestamp.
            - ``USUARIO`` (str): Identificador del usuario que ejecuta la carga.
            - ``ACTIVO`` (int): Estado lógico del registro 0/1.

    Returns:
        None

    Raises:
        requests.exceptions.RequestException: Si ocurre un error durante la
        comunicación con la API.
        KeyError: Si algún registro no contiene las claves esperadas.
        IndexError: Si la lista recibida está vacía.
    """

    print("#################### updateDeterminantes ####################")
    print("data recibida", determinantes_data)

    database_name = determinantes_data[0]["DBEVENT"]
    print("Update determinantes evento+-+-+-+-: ", database_name)

    # Obtiene los registros existentes asociados al evento actual
    endpoint = f"http://{host}:5000/api/get/{database_name}/determinantes/all/-/-/-/-/-"
    existing_determinantes = requests.get(endpoint).json()
    
    # Inicializa la estructura esperada si la consulta no retorna módulos
    if not("MODULO" in existing_determinantes):
        existing_determinantes["MODULO"] = []

    for record in determinantes_data:
        try:
            # Inserta el módulo si aún no existe en la tabla determinantes
            if not(record["MODULO"] in existing_determinantes["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/determinantes"
                post_response = requests.post(endpoint, data=json.dumps(record))
            # Actualiza el registro existente utilizando el ID asociado al módulo
            else:
                module_index = existing_determinantes["MODULO"].index(record["MODULO"])
                record_id = existing_determinantes["ID"][module_index]
                endpoint = f"http://{host}:5000/api/update/determinantes/{record_id}"
                update_response = requests.post(endpoint, data=json.dumps(record))
        except Exception as error:
            print(f"Error procesando el modulo {record.get('MODULO', 'Desconocido')}: {error}")

if __name__ == '__main__':
    #print("finished")
    refreshModules()
    makeModularities()
    #data = makeModules()
    #visionUpdate(data)
    #pdcrVariants("dumie")